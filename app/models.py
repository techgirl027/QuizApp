import os
from django.shortcuts import get_object_or_404
import openpyxl
from django.db import models
from django.contrib.auth.models import User


class Quiz(models.Model):
    name = models.CharField(max_length=100)
    author = models.ForeignKey(User, on_delete=models.CASCADE)
    amount = models.IntegerField()

    def __str__(self):
        return self.name

    @property
    def questions(self):
        return Question.objects.filter(quiz=self)

    @property
    def count_followers(self):
        return Answer.objects.filter(quiz=self).count

    @property
    def questions_count(self):
        return Question.objects.filter(quiz=self).count()

    @property
    def quiz(self):
        return Question.objects.filter(quiz=self)


class Question(models.Model):
    name = models.CharField(max_length=255)
    quiz = models.ForeignKey(Quiz, on_delete=models.CASCADE)

    def __str__(self):
        return self.name

    @property
    def options(self):
        return Option.objects.filter(question=self).order_by("?")

    @property
    def options_count(self):
        return Option.objects.filter(question=self).count()

    @property
    def correct_option(self):
        return Option.objects.get(question=self, correct=True)

    @property
    def correct_option_count(self):
        return Option.objects.filter(question=self, correct=True)

    @property
    def uncorrect_option_count(self):
        return self.options_count - self.correct_option_count


class Option(models.Model):
    name = models.CharField(max_length=100)
    question = models.ForeignKey(Question, on_delete=models.CASCADE)
    correct = models.BooleanField(default=False)

    def __str__(self):
        return self.name

    def save(self, *args, **kwargs):
        if not Option.objects.filter(question=self.question).count():  # 0,1,2,3 ...
            assert self.correct, "Birinchi javobingiz to'g'ri bo'lishi kerak"
        else:
            assert not self.correct, "Bu savolda to'g'ri javob bor"
        super(Option, self).save(*args, **kwargs)


class Answer(models.Model):
    quiz = models.ForeignKey(Quiz, on_delete=models.CASCADE)
    author = models.ForeignKey(User, on_delete=models.CASCADE)
    start_time = models.DateTimeField(null=True, blank=True)
    end_time = models.DateTimeField(null=True, blank=True)
    is_late = models.BooleanField(null=True, blank=True)

    @property
    def correct_options_count(self):
        count = 0
        for answer in AnswerDetail.objects.filter(answer=self):
            if answer.is_correct:
                count += 1
        return count

    @property
    def questions(self):
        return Question.objects.filter(quiz=self.quiz)

    @property
    def wrong_options_count(self):
        return self.quiz.questions_count - self.correct_options_count

    def __str__(self):
        return f"{self.author.username} -> {self.quiz.name}"

    def save(self, *args, **kwargs):
        super(Answer, self).save(*args, **kwargs)

    @staticmethod
    def export_to_excel(answer_id):
        answer = get_object_or_404(Answer, id=answer_id)

        answer_details = AnswerDetail.objects.filter(answer=answer)

        book = openpyxl.Workbook()
        sheet = book.active

        headers = ["Question", "User Choice", "Is Correct"]
        for col_num, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col_num).value = header

        for row_num, detail in enumerate(answer_details, start=2):
            sheet.cell(row=row_num, column=1).value = detail.question.name
            sheet.cell(row=row_num, column=2).value = detail.user_choice.name
            sheet.cell(row=row_num, column=3).value = (
                "Yes" if detail.is_correct else "No"
            )

            file_path = f"answer_{answer_id}_details.xlsx"
        book.save(file_path)

        return file_path


class AnswerDetail(models.Model):
    answer = models.ForeignKey(Answer, on_delete=models.CASCADE)
    question = models.ForeignKey(Question, on_delete=models.CASCADE)
    user_choice = models.ForeignKey(Option, on_delete=models.CASCADE)

    def save(self, *args, **kwargs):
        assert not AnswerDetail.objects.filter(
            answer=self.answer, question=self.question
        ).count(), "Bu savolga javob berilgan"
        super(AnswerDetail, self).save(*args, **kwargs)

    @property
    def is_correct(self):
        return self.user_choice == self.question.correct_option
