import os
import openpyxl
from django.shortcuts import render, redirect
from django.contrib.auth.models import User
from django.contrib.auth import login, authenticate
from django.contrib import messages
from . import models
from random import choice, sample
from django.http import HttpResponse

# Create your views here.


def home_page(request):
    context = {}
    return render(request, "index.html", context)


def register(request):
    if request.method == "POST":
        full_name = request.POST.get("full_name")
        username = request.POST.get("username")
        password1 = request.POST.get("password1")
        password2 = request.POST.get("password2")

        if password1 != password2:
            messages.error(request, "Passwords do not match.")
            return redirect("signup")

        if User.objects.filter(username=username).exists():
            messages.error(request, "Username already taken.")
            return redirect("signup")

        user = User.objects.create_user(username=username, password=password1)
        user.first_name = full_name
        user.save()

        authenticated_user = authenticate(username=username, password=password1)
        if authenticated_user is not None:
            login(request, authenticated_user)
        return redirect("index")
    return render(request, "reg.html")


def login_(request):
    if request.method == "POST":
        username = request.POST["username"]
        password = request.POST["password"]
        user = authenticate(username=username, password=password)
        if user:
            login(request, user)
            return redirect("index")
        else:
            return redirect("error")
    return render(request, "login.html")


def create_quiz(request):
    context = {}
    return render(request, "quiz.html", context)


def index(request):
    if not request.user.is_authenticated:
        return redirect("login")
    return render(request, "index.html")


def allQuizList(request):
    images = [
        "https://st2.depositphotos.com/2769299/7314/i/450/depositphotos_73146775-stock-photo-a-stack-of-books-on.jpg",
        "https://img.freepik.com/free-photo/creative-composition-world-book-day_23-2148883765.jpg",
        "https://profit.pakistantoday.com.pk/wp-content/uploads/2018/04/Stack-of-books-great-education.jpg",
        "https://live-production.wcms.abc-cdn.net.au/73419a11ea13b52c6bd9c0a69c10964e?impolicy=wcms_crop_resize&cropH=1080&cropW=1918&xPos=1&yPos=0&width=862&height=485",
        "https://live-production.wcms.abc-cdn.net.au/398836216839841241467590824c5cf1?impolicy=wcms_crop_resize&cropH=2813&cropW=5000&xPos=0&yPos=0&width=862&height=485",
        "https://images.theconversation.com/files/45159/original/rptgtpxd-1396254731.jpg?ixlib=rb-4.1.0&q=45&auto=format&w=1356&h=668&fit=crop",
    ]

    quizes = models.Quiz.objects.all()
    # images = sample(len(quizes), images)

    quizes_list = []

    for quiz in quizes:
        quiz.img = choice(images)
        quizes_list.append(quiz)

    return render(request, "all-quiz-list.html", {"quizes": quizes_list})


def quizList(request):
    images = [
        "https://st2.depositphotos.com/2769299/7314/i/450/depositphotos_73146775-stock-photo-a-stack-of-books-on.jpg",
        "https://img.freepik.com/free-photo/creative-composition-world-book-day_23-2148883765.jpg",
        "https://profit.pakistantoday.com.pk/wp-content/uploads/2018/04/Stack-of-books-great-education.jpg",
        "https://live-production.wcms.abc-cdn.net.au/73419a11ea13b52c6bd9c0a69c10964e?impolicy=wcms_crop_resize&cropH=1080&cropW=1918&xPos=1&yPos=0&width=862&height=485",
        "https://live-production.wcms.abc-cdn.net.au/398836216839841241467590824c5cf1?impolicy=wcms_crop_resize&cropH=2813&cropW=5000&xPos=0&yPos=0&width=862&height=485",
        "https://images.theconversation.com/files/45159/original/rptgtpxd-1396254731.jpg?ixlib=rb-4.1.0&q=45&auto=format&w=1356&h=668&fit=crop",
    ]

    quizes = models.Quiz.objects.filter(author=request.user)
    # images = sample(len(quizes), images)

    quizes_list = []

    for quiz in quizes:
        quiz.img = choice(images)
        quizes_list.append(quiz)

    return render(request, "quiz-list.html", {"quizes": quizes_list})


def quizDetail(request, id):
    quiz = models.Quiz.objects.get(id=id)
    return render(request, "quiz-detail.html", {"quiz": quiz})


def quizListDetail(request, id):
    quiz = models.Quiz.objects.get(id=id)
    answers = models.Answer.objects.filter(quiz=quiz)
    return render(request, "quiz-list-detail.html", {"answers": answers, "id": id})


def quizListDetailExcel(request, id):
    quiz = models.Quiz.objects.get(id=id)
    answer = models.Answer.objects.get(quiz=quiz)
    answers = models.AnswerDetail.objects.filter(answer=answer)
    book = openpyxl.Workbook()
    sheet = book.active

    # Birinchi qator: sarlavhalar
    headers = ["Username", "to'g'ri javoblar", "xato javoblar"]
    for col_num, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col_num).value = header

    # Ma'lumotlarni yozish
    for row_num, detail in enumerate(answers, start=2):
        sheet.cell(row=row_num, column=1).value = detail.answer.author.username
        sheet.cell(row=row_num, column=2).value = detail.answer.correct_options_count
        sheet.cell(row=row_num, column=3).value = detail.answer.wrong_options_count

    file_path = f"answer__details.xlsx"
    book.save(file_path)

    # Faylni ochib, HTTP responsega yuborish
    with open(file_path, "rb") as f:
        response = HttpResponse(
            f.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = (
            f'attachment; filename="{os.path.basename(file_path)}"'
        )

    # Yaratilgan faylni o'chirish
    os.remove(file_path)

    return response


def questionDelete(request, id, pk):
    models.Question.objects.get(id=id).delete()
    return redirect("quizDetail", id=pk)


def createQuiz(request):
    if request.method == "POST":
        quiz = models.Quiz.objects.create(
            name=request.POST["name"],
            amount=request.POST["amount"],
            author=request.user,
        )
        return redirect("quizDetail", quiz.id)
    return render(request, "quiz-create.html")


def questionCreate(request, id):
    quiz = models.Quiz.objects.get(id=id)
    if request.method == "POST":
        question_text = request.POST["name"]
        true = request.POST["true"]
        false_list = request.POST.getlist("false-list")

        question = models.Question.objects.create(
            name=question_text,
            quiz=quiz,
        )
        question.save()
        models.Option.objects.create(
            question=question,
            name=true,
            correct=True,
        )

        for false in false_list:
            models.Option.objects.create(
                name=false,
                question=question,
            )
        return redirect("quizList")

    return render(request, "question-create.html", {"quiz": quiz})


def questionDetail(request, id):
    question = models.Question.objects.get(id=id)
    return render(request, "question-detail.html", {"question": question})


def deleteOption(request, ques, option):
    question = models.Question.objects.get(id=ques)
    models.Option.objects.get(question=question, id=option).delete()
    return redirect("questionDetail", id=ques)


def answerDetail(request, id):
    answer = models.Answer.objects.get(id=id)
    answers = models.AnswerDetail.objects.filter(answer=answer)
    return render(request, "answer-detail.html", {"result": answers})


def answerDetailExcel(request, id):
    quiz = models.Quiz.objects.get(id=id)
    answer = models.Answer.objects.get(quiz=quiz, author=request.user)
    answers = models.AnswerDetail.objects.filter(answer=answer)
    book = openpyxl.Workbook()
    sheet = book.active

    headers = ["Savol", "javob", "berilgan javob", "natija"]
    for col_num, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col_num).value = header

    for row_num, detail in enumerate(answers, start=2):
        sheet.cell(row=row_num, column=1).value = detail.question.name
        sheet.cell(row=row_num, column=2).value = detail.question.correct_option.name
        sheet.cell(row=row_num, column=3).value = detail.user_choice.name
        sheet.cell(row=row_num, column=4).value = detail.is_correct

    file_path = f"answer_detail.xlsx"
    book.save(file_path)

    with open(file_path, "rb") as f:
        response = HttpResponse(
            f.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = (
            f'attachment; filename="{os.path.basename(file_path)}"'
        )

    os.remove(file_path)
    return response
