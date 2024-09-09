import os
import openpyxl
from django.http import HttpResponse
from django.shortcuts import render, redirect
from main import models


def getQuiz(request, id):
    quiz = models.Quiz.objects.get(id=id)
    return render(request, "answer/get-quiz.html", {"quiz": quiz})


def quizDetail(request, id):
    quiz = models.Quiz.objects.get(id=id)
    return render(request, "answer/quiz-detail.html", {"quiz": quiz})


def makeAnswer(request, id):
    quiz = models.Quiz.objects.get(id=id)
    is_have = models.Answer.objects.get(quiz=quiz, author=request.user)
    if is_have:
        return redirect("quiz_detail", id)
    answer = models.Answer.objects.create(quiz=quiz, author=request.user)
    for key, value in request.POST.items():
        if key.isdigit():
            models.AnswerDetail.objects.create(
                answer=answer,
                question=models.Question.objects.get(id=int(key)),
                user_choice=models.Option.objects.get(id=int(value)),
            )
    return redirect("results")


def results(request):
    results = models.Answer.objects.filter(author=request.user)
    return render(request, "answer/natijalar.html", {"result_list": results})


def result_detail(request, id):
    quiz = models.Quiz.objects.get(id=id)
    answer = models.Answer.objects.get(quiz=quiz, author=request.user)
    answers = models.AnswerDetail.objects.filter(answer=answer)
    return render(request, "answer/natija-detail.html", {"result": answers, "id": id})


def answerDetailExcel(request, id):
    quiz = models.Quiz.objects.get(id=id)
    answer = models.Answer.objects.get(quiz=quiz, author=request.user)
    answers = models.AnswerDetail.objects.filter(answer=answer)
    book = openpyxl.Workbook()
    sheet = book.active

    # Birinchi qator: sarlavhalar
    headers = ["Savol", "javob", "berilgan javob", "natija"]
    for col_num, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col_num).value = header

    # Ma'lumotlarni yozish
    for row_num, detail in enumerate(answers, start=2):
        sheet.cell(row=row_num, column=1).value = detail.question.name
        sheet.cell(row=row_num, column=2).value = detail.question.correct_option.name
        sheet.cell(row=row_num, column=3).value = detail.user_choice.name
        sheet.cell(row=row_num, column=4).value = detail.is_correct

    file_path = f"answer__details.xlsx"
    book.save(file_path)

    # Faylni ochib, HTTP responsega yuborish
    with open(file_path, "rb") as f:
        response = HttpResponse(
            f.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = (
            f'attachment; filename="{os.path.basename(file_path)}"'
        )

    # Yaratilgan faylni o'chirish
    os.remove(file_path)

    return response
